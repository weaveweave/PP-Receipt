"""
Step 3: Convert extracted JSON data to Excel spreadsheet
Creates a nicely formatted Excel file with all receipt data
Updated to handle new format with multiple receipts per image
FIXED: no_kuitansi now treated as string, not numeric
"""

import os
import json
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

from config import (
    STEP2_OUTPUT_DIR,
    STEP3_OUTPUT_DIR,
    EXCEL_OUTPUT_FILENAME,
    EXCEL_SHEET_NAME,
    EXCEL_COLUMNS,
    EXCEL_AUTO_WIDTH,
    EXCEL_FREEZE_HEADER,
    EXCEL_ADD_BORDERS,
    EXCEL_HEADER_BOLD,
    VERBOSE
)


# ============================================================================
# JSON LOADING
# ============================================================================

def load_all_json_files(json_dir):
    """
    Load all JSON files from directory and extract receipts
    Now handles the new format with receipts array
    
    Args:
        json_dir (str): Directory containing JSON files
    
    Returns:
        list: List of dictionaries with extracted data (flattened from all receipts arrays)
    """
    
    json_files = [
        os.path.join(json_dir, f)
        for f in os.listdir(json_dir)
        if f.endswith('.json')
    ]
    
    if not json_files:
        print("⚠️  No JSON files found!")
        return []
    
    print(f"📊 Found {len(json_files)} JSON file(s)")
    
    all_data = []
    errors = []
    total_receipts = 0
    
    for json_path in sorted(json_files):
        filename = Path(json_path).name
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                
                # Handle new format with receipts array
                if isinstance(data, dict) and "receipts" in data:
                    receipts = data["receipts"]
                    
                    # Case 1: receipts is a list
                    if isinstance(receipts, list):
                        for receipt in receipts:
                            if receipt and isinstance(receipt, dict):  # Skip empty or invalid receipts
                                all_data.append(receipt)
                                total_receipts += 1
                        if VERBOSE:
                            print(f"  ✓ Loaded: {filename} ({len(receipts)} receipt(s))")
                    
                    # Case 2: receipts is a single dict (LLM mistake)
                    elif isinstance(receipts, dict):
                        # Treat it as a single receipt
                        all_data.append(receipts)
                        total_receipts += 1
                        if VERBOSE:
                            print(f"  ✓ Loaded: {filename} (1 receipt - dict format)")
                    
                    # Case 3: receipts is a STRING (needs parsing)
                    elif isinstance(receipts, str):
                        try:
                            # Try to parse as JSON first
                            parsed = json.loads(receipts)
                            if isinstance(parsed, list):
                                for receipt in parsed:
                                    if receipt and isinstance(receipt, dict):
                                        all_data.append(receipt)
                                        total_receipts += 1
                                if VERBOSE:
                                    print(f"  ✓ Loaded: {filename} ({len(parsed)} receipt(s) - parsed from string)")
                            elif isinstance(parsed, dict):
                                all_data.append(parsed)
                                total_receipts += 1
                                if VERBOSE:
                                    print(f"  ✓ Loaded: {filename} (1 receipt - parsed from string)")
                        except json.JSONDecodeError:
                            # Try to parse as Python literal (single quotes)
                            try:
                                import ast
                                parsed = ast.literal_eval(receipts)
                                if isinstance(parsed, list):
                                    for receipt in parsed:
                                        if receipt and isinstance(receipt, dict):
                                            all_data.append(receipt)
                                            total_receipts += 1
                                    if VERBOSE:
                                        print(f"  ✓ Loaded: {filename} ({len(parsed)} receipt(s) - parsed Python literal)")
                                elif isinstance(parsed, dict):
                                    all_data.append(parsed)
                                    total_receipts += 1
                                    if VERBOSE:
                                        print(f"  ✓ Loaded: {filename} (1 receipt - parsed Python literal)")
                            except (ValueError, SyntaxError) as e:
                                errors.append({"file": filename, "error": f"Could not parse string receipts: {e}"})
                                print(f"  ⚠️  Warning: {filename} - Could not parse receipts string, skipping")
                    
                    # Case 4: receipts is something else (number, etc)
                    else:
                        # Try to salvage: check if data itself has receipt fields
                        required_fields = ["no_kuitansi", "tanggal", "penerima"]
                        if all(field in data for field in required_fields):
                            all_data.append(data)
                            total_receipts += 1
                            if VERBOSE:
                                print(f"  ✓ Loaded: {filename} (1 receipt - salvaged from root)")
                        else:
                            errors.append({"file": filename, "error": f"receipts field is {type(receipts).__name__}, not list"})
                            print(f"  ⚠️  Warning: {filename} - receipts field is {type(receipts).__name__}, skipping")
                
                # Handle old format (single receipt per file, no receipts field)
                elif isinstance(data, dict):
                    # Check if it has receipt fields
                    required_fields = ["no_kuitansi", "tanggal", "penerima"]
                    if any(field in data for field in required_fields):
                        all_data.append(data)
                        total_receipts += 1
                        if VERBOSE:
                            print(f"  ✓ Loaded: {filename} (1 receipt - old format)")
                    else:
                        errors.append({"file": filename, "error": "No receipt fields found"})
                        print(f"  ⚠️  Warning: {filename} - No valid receipt fields found")
                
                # Handle if data is a list directly
                elif isinstance(data, list):
                    for item in data:
                        if isinstance(item, dict):
                            all_data.append(item)
                            total_receipts += 1
                    if VERBOSE:
                        print(f"  ✓ Loaded: {filename} ({len(data)} receipt(s) - list format)")
                
                else:
                    errors.append({"file": filename, "error": "Invalid JSON structure"})
                    print(f"  ✗ Error: {filename} - Invalid JSON structure")
                    
        except Exception as e:
            errors.append({"file": filename, "error": str(e)})
            print(f"  ✗ Error loading {filename}: {e}")
    
    print(f"✅ Successfully loaded {total_receipts} receipt(s) from {len(json_files)} file(s)")
    
    if errors:
        print(f"⚠️  {len(errors)} file(s) had issues (but may have been salvaged)")
    
    return all_data


# ============================================================================
# DATA TRANSFORMATION
# ============================================================================

def transform_to_dataframe(data_list):
    """
    Transform list of JSON data to pandas DataFrame
    
    Args:
        data_list (list): List of dictionaries
    
    Returns:
        pd.DataFrame: DataFrame with standardized columns
    """
    
    if not data_list:
        print("⚠️  No data to transform!")
        return pd.DataFrame()
    
    # Create DataFrame
    df = pd.DataFrame(data_list)
    
    # Ensure all required columns exist
    required_fields = ["no_kuitansi", "tanggal", "penerima", "uang_sejumlah_rp", "jumlah_liter", "keterangan"]
    for field in required_fields:
        if field not in df.columns:
            df[field] = ""
    
    # Reorder columns to match EXCEL_COLUMNS
    df = df[required_fields]
    
    # Rename columns to display names
    df.columns = EXCEL_COLUMNS
    
    # Format currency column (add thousands separator)
    currency_col = "Uang Sejumlah (Rp)"
    if currency_col in df.columns:
        df[currency_col] = df[currency_col].apply(lambda x: format_currency(x))
    
    # FIXED: Keep no_kuitansi as string, don't convert to numeric
    if "No. Kuitansi" in df.columns:
        # Convert to string and handle any None/NaN values
        df["No. Kuitansi"] = df["No. Kuitansi"].fillna("").astype(str)
        # Remove any ".0" that might have been added if value was originally numeric
        df["No. Kuitansi"] = df["No. Kuitansi"].str.replace(".0", "", regex=False)
        
        # Sort by No. Kuitansi as string (alphanumeric sort)
        # This will sort: "001", "002", "010", "A01", "B02", etc.
        df = df.sort_values("No. Kuitansi")
    
    print(f"📊 Created DataFrame with {len(df)} rows")
    
    return df


def format_currency(value):
    """
    Format currency value with thousands separator
    
    Args:
        value: Currency value (string or number)
    
    Returns:
        str: Formatted currency string
    """
    
    if not value or value == "":
        return ""
    
    try:
        # Convert to integer
        num = int(str(value).replace(",", "").replace(".", ""))
        # Format with thousands separator
        return f"{num:,}".replace(",", ".")
    except:
        return str(value)


# ============================================================================
# EXCEL FORMATTING
# ============================================================================

def apply_excel_formatting(excel_path):
    """
    Apply formatting to Excel file: bold headers, borders, auto-width, etc.
    
    Args:
        excel_path (str): Path to Excel file
    """
    
    print("🎨 Applying Excel formatting...")
    
    # Load workbook
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # Define styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    # Apply header formatting
    if EXCEL_HEADER_BOLD:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            if EXCEL_ADD_BORDERS:
                cell.border = thin_border
    
    # Apply borders to all cells
    if EXCEL_ADD_BORDERS:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                # Center-align numeric columns
                if cell.column in [1, 4, 5]:  # No. Kuitansi, Uang Sejumlah, Jumlah Liter
                    cell.alignment = center_alignment
                else:
                    cell.alignment = left_alignment
    
    # Auto-adjust column widths
    if EXCEL_AUTO_WIDTH:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            # Set width with some padding
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    if EXCEL_FREEZE_HEADER:
        ws.freeze_panes = "A2"
    
    # Save workbook
    wb.save(excel_path)
    print("✅ Formatting applied")


# ============================================================================
# EXCEL GENERATION
# ============================================================================

def create_excel(data_list, output_path):
    """
    Create Excel file from JSON data
    
    Args:
        data_list (list): List of dictionaries with receipt data
        output_path (str): Path to save Excel file
    
    Returns:
        bool: True if successful, False otherwise
    """
    
    if not data_list:
        print("❌ No data to create Excel file!")
        return False
    
    try:
        # Transform to DataFrame
        df = transform_to_dataframe(data_list)
        
        if df.empty:
            print("❌ DataFrame is empty!")
            return False
        
        # Save to Excel
        print(f"💾 Saving to Excel: {output_path}")
        df.to_excel(output_path, sheet_name=EXCEL_SHEET_NAME, index=False, engine='openpyxl')
        
        # Apply formatting
        apply_excel_formatting(output_path)
        
        print(f"✅ Excel file created successfully!")
        print(f"   - Rows: {len(df)}")
        print(f"   - Columns: {len(df.columns)}")
        
        return True
    
    except Exception as e:
        print(f"❌ Error creating Excel file: {e}")
        return False


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function"""
    
    print("=" * 60)
    print("🚀 STEP 3: CONVERT JSON TO EXCEL")
    print("=" * 60)
    print(f"Input directory: {STEP2_OUTPUT_DIR}")
    print(f"Output directory: {STEP3_OUTPUT_DIR}")
    print(f"Output filename: {EXCEL_OUTPUT_FILENAME}")
    
    # Check if input directory exists
    if not os.path.exists(STEP2_OUTPUT_DIR):
        print(f"\n❌ Error: Input directory not found: {STEP2_OUTPUT_DIR}")
        print("   Please run step2_extract_data.py first")
        return
    
    # Load all JSON files
    print(f"\n📂 Loading JSON files from {STEP2_OUTPUT_DIR}...")
    data_list = load_all_json_files(STEP2_OUTPUT_DIR)
    
    if not data_list:
        print("\n❌ No data found to process!")
        return
    
    # Create output directory
    os.makedirs(STEP3_OUTPUT_DIR, exist_ok=True)
    
    # Create Excel file
    output_path = os.path.join(STEP3_OUTPUT_DIR, EXCEL_OUTPUT_FILENAME)
    print(f"\n📊 Creating Excel spreadsheet...")
    
    success = create_excel(data_list, output_path)
    
    if success:
        print("\n✨ Step 3 completed!")
        print(f"📁 Output saved to: {output_path}")
        print(f"\n🎉 Pipeline completed successfully!")
        print(f"   You can now open the Excel file to view all extracted receipt data.")
    else:
        print("\n❌ Failed to create Excel file")


if __name__ == "__main__":
    main()
